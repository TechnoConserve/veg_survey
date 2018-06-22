import argparse
import bisect

from openpyxl import load_workbook

description = """For summary and analysis of a vegetation survey file. This program makes the following assumptions:

1. Site data is split among sheets in the the file to be processed.
2. The name of the site is contained in column B, row 2 of each sheet.
"""
parser = argparse.ArgumentParser(description=description)
parser.add_argument("-d", "--data", action="store", required=True, dest="data", help="Path to your data file")

args = parser.parse_args()
data = args.data

VARIABLE_LIST = ['DNF', 'DF', 'FU']
NUM_PHENOLOGIES = 2


class Species:
    """
    An object describing the density characteristics of a particular
    species during a vegetation survey.
    """
    def __init__(self, name, non_flowering_density, flowering_density, flowering_units):
        self.id = name
        if non_flowering_density is None:
            self.dnf = 0
        else:
            self.dnf = non_flowering_density
        if flowering_density is None:
            self.df = 0
        else:
            self.df = flowering_density
        if flowering_units is None:
            self.fu = 0
        else:
            self.fu = flowering_units


class Survey:
    """
    An object describing a vegetation survey. Each object is identified
    by a site name, a point in time, and a list of species objects.
    """
    all_surveys = []
    survey_dates = []

    def __init__(self, site, date, species_comp):
        self.site = site
        self.date = date
        self.species = species_comp

        Survey._calc_num_phenologies(self)
        self.all_surveys.append(self)

    def get_dates_site_surveyed(self):
        """
        Get a list of dates this site was surveyed.
        :return: A list of datetime objects this site was surveyed.
        """
        return [other_survey.date for other_survey in Survey.all_surveys if self.site == other_survey.site]

    def get_other_surveys_from_site(self):
        """
        Get other surveys conducted at this site.
        :return: A list of Survey objects with the same site name.
        """
        return [other_survey for other_survey in Survey.all_surveys if self.site == other_survey.site]

    @classmethod
    def _calc_num_phenologies(cls, survey):
        """
        Determine what phenology number this survey represents based on
        other surveys conducted at the same site. Update phen_num of
        other surveys and add the date to the list of survey dates if
        it is not already there.
        :param survey: Survey object.
        :return: None.
        """
        other_surveys = survey.get_other_surveys_from_site()
        dates_surveyed = survey.get_dates_site_surveyed()
        bisect.insort(dates_surveyed, survey.date)
        survey.phen_num = dates_surveyed.index(survey.date) + 1
        surveys_before = [earlier_survey for earlier_survey in other_surveys if earlier_survey.date < survey.date]
        surveys_after = [later_survey for later_survey in other_surveys if later_survey.date > survey.date]
        for surv in surveys_before:
            surv.phen_num -= 1
        for surv in surveys_after:
            surv.phen_num += 1

        if survey.date not in cls.survey_dates:
            bisect.insort(cls.survey_dates, survey.date)

    @classmethod
    def get_all_surveys(cls):
        return cls.all_surveys

    @classmethod
    def get_survey_dates(cls):
        return cls.survey_dates


def get_data(file_loc):
    """
    Attempt to load the specified file into an openpyxl Workbook object.
    :return: An openpyxl Workbook
    """
    print("Attempting to open data file...")
    try:
        # Use the data_only flag so we get values computed by formulae
        wb = load_workbook(file_loc, data_only=True)
        print("File loaded!")
        return wb
    except FileNotFoundError as err:
        print("[!] Error Message: {}".format(err))
        print("Failed to find file. Did you specify the correct file path?")


def get_sheets(wb):
    """
    Read the given Workbook and return the list of sheet names.
    :param wb: openpyxl Workbook object
    :return: A list of sheet names.
    """
    sheets = wb.get_sheet_names()
    print("The following sheets are in this workbook:", sheets)
    return sheets


def get_site_name(sheet):
    """
    Assuming that the name of the survey exists in cell B2 of the
    given sheet, returns the name value associated with the survey
    data in the sheet.
    :param sheet: An openpyxl Sheet object.
    :return: A string value representing the name of the survey site.
    """
    site_name = sheet['B2'].value
    return site_name


def get_survey_date(sheet):
    """
    Assuming that the name of the survey exists in cell D2 of the
    given sheet, returns datetime value of the date associated with
    the survey data in the sheet.
    :param sheet: An openpyxl Sheet object.
    :return: A datetime object representing the date the survey was
    conducted.
    """
    site_date = sheet['D2'].value
    return site_date


def get_survey_species(sheet):
    species_list = []
    species_rows = sheet.rows
    # Skip the rows that are just headers
    [next(species_rows) for _ in range(3)]
    for row in species_rows:
        # Get the species name from the first column
        species_name = row[0].value
        # Get the non-flowering density from the second column
        dnf = row[1].value
        # Get the flowering density from the third column
        df = row[2].value
        # Get the number for floral units from the fourth column
        fu = row[3].value
        species = Species(species_name, dnf, df, fu)
        species_list.append(species)
        print("{} added to species composition list".format(species_name))

    return species_list


def write_summary_headers(sum_ws, all_species):
    sum_ws['A1'] = "Site"
    sum_ws['B1'] = "Variable"
    sum_ws['C1'] = "Phenology"

    col_gen = sum_ws.iter_cols(min_col=4, max_col=len(all_species) + 3, max_row=1)

    # Write species column header for each species
    for species in all_species:
        try:
            cell = next(col_gen)[0]
            cell.value = species
        except StopIteration:
            print("[!] Failed to write header for {}!".format(species))


def write_sites_to_rows(sum_ws, site_names):
    name_copy = site_names.copy()
    # Repeat site names based on number of variables
    name_copy *= len(VARIABLE_LIST) * NUM_PHENOLOGIES
    for cell, site_name in zip(sum_ws.iter_rows(min_row=2, max_row=len(name_copy) + 1, max_col=1), name_copy):
        cell[0].value = site_name


def write_variables_to_rows(sum_ws, num_sites):
    # Repeat variable names based on number of sites, variables & phenologies
    variables = sorted(VARIABLE_LIST * num_sites * NUM_PHENOLOGIES)
    for cell, variable in zip(sum_ws.iter_rows(min_row=2, max_row=len(variables) + 1, min_col=2, max_col=3), variables):
        cell[0].value = variable


def write_phenology_to_rows(sum_ws, num_sites):
    # Create list of phenologies based on number of sites
    phenology_list = sorted(list(range(1, NUM_PHENOLOGIES + 1)) * num_sites) * len(VARIABLE_LIST)
    for cell, phenology in zip(sum_ws.iter_rows(min_row=2, max_row=len(phenology_list) + 1, min_col=3, max_col=4),
                               phenology_list):
        cell[0].value = phenology


def write_data_to_rows(sum_ws, survey_list, num_sites, species_list):
    # Calculate number of rows
    num_rows = num_sites * NUM_PHENOLOGIES * len(VARIABLE_LIST)
    for row in sum_ws.iter_rows(min_row=2, max_row=num_rows + 1, max_col=len(species_list) + 3):
        site_name = row[0].value
        variable = row[1].value
        phenology = row[2].value

        for idx, species_name in enumerate(species_list):
            survey = [survey for survey in survey_list
                      if survey.site == site_name and survey.phen_num == phenology]
            if len(survey) == 1:
                survey = survey[0]
                species_obj = [species for species in survey.species if species.id == species_name]
                if len(species_obj) == 1:
                    species_obj = species_obj[0]
                    if variable == "DNF":
                        datum = species_obj.dnf
                    elif variable == "DF":
                        datum = species_obj.df
                    elif variable == "FU":
                        datum = species_obj.fu
                    else:
                        print("[!] Unrecognized variable!", variable)
                        datum = None

                    row[idx + 3].value = datum

                elif len(species_obj) > 1:
                    print("[!] Unexpected number of matching species! Expected 1, got {}: {}"
                          .format(len(species_obj), species_obj))
                else:
                    pass  # Expected since not all surveys have all species

            elif len(survey) > 1:
                print("[!] Unexpected number of matching surveys! Expected 1, got {}: {}".format(len(survey), survey))
            else:
                print("[!] No matching Surveys found!")


def main():
    wb = get_data(data)
    all_species = []
    all_sites = []
    
    # Start processing the data
    for sheet in wb:
        site_name = get_site_name(sheet)
        all_sites.append(site_name)
        survey_date = get_survey_date(sheet)
        print("Working on site: {} ({:%Y-%m-%d})\n".format(site_name, survey_date))
        species_comp = get_survey_species(sheet)
        all_species += species_comp
        # Create a survey object from the data contained in the sheet
        Survey(site=site_name, date=survey_date, species_comp=species_comp)

    # Summarize the data
    species_ids = []
    for species in all_species:
        species_ids.append(species.id)
    species_ids = sorted(set(species_ids))

    all_sites = sorted(set(all_sites))

    print("\n[*] Summarizing data...")
    sum_ws = wb.create_sheet("Summary")
    num_sites = len(all_sites)
    write_summary_headers(sum_ws, species_ids)
    write_sites_to_rows(sum_ws, all_sites)
    write_variables_to_rows(sum_ws, num_sites)
    write_phenology_to_rows(sum_ws, num_sites)
    write_data_to_rows(sum_ws, Survey.get_all_surveys(), num_sites, species_ids)

    # Save a copy of the data with the summary
    file_name, extension = data.split('.')
    file_name = file_name + "_summarized" + "." + extension
    wb.save(file_name)
    print("[*] Saved summary file to {}!".format(file_name))


if __name__ == "__main__":
    main()
